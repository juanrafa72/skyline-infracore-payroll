"""
Exporta el histórico de los dos libros de nómina a JSON, sin interpretar nada.

Solo lee y normaliza formatos (fechas, mayúsculas, espacios). NO decide si dos
nombres son la misma persona, ni si un nombre es una máquina o una cuadrilla:
esas decisiones las toma un humano en la aplicación.

Los Excel originales NO se modifican.

Uso:  ./venv/bin/python export_history.py salida.json
"""
import json
import sys
from datetime import datetime

import openpyxl

SKYLINE = "/Users/jrc/Documents/Documentos - MacBook Pro de Juan/FIBRA OPTICA/SKYLINE ADVANCE TECH/03 - RECURSOS HUMANOS/SEGUIMIENTO LABORAL/2026 Laboral/NOMINA 2026 AERIO&UG- v03-20 oct---.xlsx"
INFRACORE = "/Users/jrc/Documents/Documentos - MacBook Pro de Juan/FIBRA OPTICA/INFRACORE SYSTEMS LLC/SEGUIMIENTO LABORAL - INFRACORE/NOMINA 2026 -INFRACORE-N-AGOSTO 9.xlsx"

DAY_TYPE = {"si": "FULL_DAY", "no": "NO_WORK", "0.5": "HALF_DAY"}

REASON = {
    "lluvia": "RAIN",
    "vacaciones": "VACATION",
    "mantenimiento": "MAINTENANCE",
    "descanso": "REST",
}


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def day_type_of(raw):
    """'Si' / 'SI' / 'no' / 0.5 → tipo normalizado. Devuelve None si no se entiende."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return "HALF_DAY" if float(raw) == 0.5 else None
    return DAY_TYPE.get(str(raw).strip().lower())


def iso(value):
    if isinstance(value, datetime):
        # Las fechas vacías de Excel llegan como 1899/1900: no son fechas reales.
        if value.year < 2000:
            return None
        return value.date().isoformat()
    return None


def money(value):
    if value is None:
        return None
    try:
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return None


def read_workbook(path, company_default, header_map):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["BASE"]

    header = None
    index = {}
    rows = []
    skipped = {"sin_fecha": 0, "sin_nombre": 0, "fila_total": 0, "dia_ilegible": 0}

    for position, raw in enumerate(ws.iter_rows(values_only=True)):
        if header is None:
            header = [clean(v) or "" for v in raw]
            index = {name: header.index(name) for name in header_map.values() if name in header}
            continue

        def cell(key):
            column = header_map.get(key)
            if column is None or column not in index:
                return None
            position_in_row = index[column]
            return raw[position_in_row] if position_in_row < len(raw) else None

        # Fila de totales incrustada en la tabla (el error C1 de Infracore).
        unit = clean(cell("company"))
        if unit and unit.lower() in {"total", "total general"}:
            skipped["fila_total"] += 1
            continue

        name = clean(cell("name"))
        date = iso(cell("date"))

        if not name:
            skipped["sin_nombre"] += 1
            continue
        if not date:
            skipped["sin_fecha"] += 1
            continue

        day = day_type_of(cell("day_worked"))
        if day is None:
            skipped["dia_ilegible"] += 1
            continue

        rows.append(
            {
                "sourceRow": position + 1,
                "company": (unit or company_default).upper(),
                "operation": clean(cell("operation")),
                "name": name,
                "date": date,
                "dayType": day,
                "reason": REASON.get((clean(cell("work_note")) or "").lower()),
                "additionCategory": clean(cell("addition_kind")),
                "additionAmount": money(cell("addition_amount")),
                "deductionAmount": money(cell("deduction_amount")),
                "deductionNote": clean(cell("deduction_note")),
                "expectedTotal": money(cell("total")),
                "rate": money(cell("rate")),
                "project": clean(cell("project")),
                "customer": clean(cell("customer")),
                "crew": clean(cell("crew")),
                "note": clean(cell("note")),
            }
        )

    wb.close()
    return rows, skipped


def read_lists(path):
    """Catálogos: trabajadores con tarifa por operación, y proyectos con cliente."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Listas"]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    def column_block(first, operation):
        out = []
        for row in grid[2:]:
            if first >= len(row):
                continue
            name = clean(row[first])
            if not name:
                continue
            out.append(
                {
                    "name": name,
                    "rate": money(row[first + 1]) if first + 1 < len(row) else None,
                    "item": clean(row[first + 2]) if first + 2 < len(row) else None,
                    "operation": operation,
                }
            )
        return out

    workers = (
        column_block(5, "ADMIN")          # F
        + column_block(9, "AERIAL")       # J
        + column_block(13, "UNDERGROUND") # N
        + column_block(17, "BLOWFIBER")   # R
    )

    projects = []
    for row in grid[2:]:
        if len(row) < 28:
            continue
        name = clean(row[25])  # Z
        if not name:
            continue
        projects.append(
            {
                "name": name,
                "customer": clean(row[26]),  # AA
                "crew": clean(row[27]),      # AB
            }
        )

    return workers, projects


def main():
    destination = sys.argv[1] if len(sys.argv) > 1 else "historico.json"

    header_skyline = {
        "operation": "Tipo", "name": "Nombre", "date": "Fecha",
        "day_worked": "Día trabajados", "work_note": "Observ Trabaj",
        "addition_kind": "Adicionales", "addition_amount": "Valor adicional",
        "deduction_amount": "Dctos", "deduction_note": "Observación Dctos",
        "total": "PAGO TOTAL", "rate": "Valor Día",
        "project": "PROYECTOS", "customer": "EMPRESA", "crew": "EQUIPO",
    }
    header_infracore = dict(header_skyline)
    header_infracore.update({"company": "UNIDAD DE NEGOCIO", "customer": "EMPRESA2", "note": "NOTA"})

    skyline_rows, skyline_skipped = read_workbook(SKYLINE, "SKYLINE", header_skyline)
    infracore_rows, infracore_skipped = read_workbook(INFRACORE, "INFRACORE", header_infracore)
    workers, projects = read_lists(SKYLINE)

    payload = {
        "generatedFrom": {"skyline": SKYLINE, "infracore": INFRACORE},
        "catalog": {"workers": workers, "projects": projects},
        "entries": skyline_rows + infracore_rows,
        "skipped": {"skyline": skyline_skipped, "infracore": infracore_skipped},
    }

    with open(destination, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)

    print(f"días exportados: {len(payload['entries'])}")
    print(f"  Skyline:   {len(skyline_rows)}  descartados: {skyline_skipped}")
    print(f"  Infracore: {len(infracore_rows)}  descartados: {infracore_skipped}")
    print(f"catálogo: {len(workers)} nombres con tarifa, {len(projects)} proyectos")
    print(f"archivo: {destination}")


if __name__ == "__main__":
    main()
