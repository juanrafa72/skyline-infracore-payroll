import sys, openpyxl
from openpyxl.utils import get_column_letter

path = sys.argv[1]
print("="*100)
print("FILE:", path)
wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
print("SHEETS (%d):" % len(wb.sheetnames))
for ws in wb.worksheets:
    print("  - %-40s state=%-8s dims=%-12s max_row=%-6s max_col=%s" % (
        ws.title, ws.sheet_state, ws.dimensions, ws.max_row, ws.max_column))
print()
for ws in wb.worksheets:
    print("-"*100)
    print("SHEET: %s  (%s rows x %s cols) state=%s" % (ws.title, ws.max_row, ws.max_column, ws.sheet_state))
    if ws.max_row is None or ws.max_row == 0:
        continue
    nrows = min(ws.max_row, 14)
    ncols = min(ws.max_column, 40)
    for r in range(1, nrows+1):
        cells = []
        for c in range(1, ncols+1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).replace("\n", "\\n")
            if len(s) > 38:
                s = s[:38] + "…"
            cells.append("%s%d=%s" % (get_column_letter(c), r, s))
        if cells:
            print("   " + " | ".join(cells))
    print()
