import sys, openpyxl
from openpyxl.utils import get_column_letter

path = sys.argv[1]
only = sys.argv[2] if len(sys.argv) > 2 else None
maxrows = int(sys.argv[3]) if len(sys.argv) > 3 else 400

wb = openpyxl.load_workbook(path, data_only=False)
wbv = openpyxl.load_workbook(path, data_only=True)

for ws in wb.worksheets:
    if only and only.lower() not in ws.title.lower():
        continue
    wsv = wbv[ws.title]
    print("="*110)
    print("SHEET: %s   (%s x %s)  state=%s" % (ws.title, ws.max_row, ws.max_column, ws.sheet_state))
    if ws.merged_cells.ranges:
        print("MERGED:", ", ".join(str(r) for r in list(ws.merged_cells.ranges)[:40]))
    print("="*110)
    for r in range(1, min(ws.max_row, maxrows)+1):
        parts = []
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            col = get_column_letter(c)
            if isinstance(v, str) and v.startswith("="):
                cv = wsv.cell(row=r, column=c).value
                parts.append("%s%d{%s}=>%s" % (col, r, v, cv))
            else:
                if isinstance(v, float):
                    v = round(v, 4)
                s = str(v).replace("\n", "\\n")
                parts.append("%s%d=%s" % (col, r, s))
        if parts:
            print(" | ".join(parts))
