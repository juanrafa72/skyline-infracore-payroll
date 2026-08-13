import sys, openpyxl
from openpyxl.utils import get_column_letter
path, only = sys.argv[1], sys.argv[2]
maxrows = int(sys.argv[3]) if len(sys.argv) > 3 else 300
startrow = int(sys.argv[4]) if len(sys.argv) > 4 else 1
data_only = (len(sys.argv) > 5 and sys.argv[5] == "v")
wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
for ws in wb.worksheets:
    if only.lower() not in ws.title.lower():
        continue
    print("=== SHEET: %s (%s x %s)" % (ws.title, ws.max_row, ws.max_column))
    r = 0
    for row in ws.iter_rows(min_row=startrow, max_row=startrow+maxrows-1):
        r += 1
        parts = []
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, float):
                v = round(v, 4)
            s = str(v).replace("\n", "\\n")
            if len(s) > 60: s = s[:60] + "…"
            parts.append("%s%s=%s" % (cell.column_letter, cell.row, s))
        if parts:
            print(" | ".join(parts))
wb.close()
